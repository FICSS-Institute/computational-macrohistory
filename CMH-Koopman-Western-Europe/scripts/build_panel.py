"""
WP-2026-006 — Panel harmonisation
==================================

Reads the three source datasets (V-Dem v16, Maddison Project 2023, COW NMC v6.0)
from raw/, joins them on (ISO3, year), and produces two outputs:

  1. processed/WP-2026-006_WesternEurope_Panel_v0-1.csv
     Long-format harmonised panel, one row per (country, year, variable).
     This is the canonical analytical dataset for §6 of the paper.

  2. WP-2026-006_DataCollection_Prospect_v0-1_filled.xlsx
     The data collection prospect with the Value column populated.
     Saved alongside the original (with _filled suffix to preserve the
     unfilled template).

Design decisions encoded:
  - Eight Western European countries: FRA, GBR, DEU, ITA, ESP, NLD, SWE, BEL
  - Two sub-windows: 1820–1913 (94 years), 1945–2020 (76 years)
  - Excluded years: 1914–1918 (WWI) and 1939–1944 (WWII active conflict).
    1945 is retained as a transition year in the post-war sub-window,
    consistent with V-Dem and Maddison practice.
  - Germany 1945–1948: flagged as "Germany under Allied occupation"
  - Germany 1949–1954: FRG existed but not COW-sovereign until 5 May 1955
    (Deutschlandvertrag); cinc unavailable for these years.
  - Germany 1955–1989: COW cinc taken from code 260 (FRG/West Germany)
  - Germany pre-1945 and 1990+: COW cinc taken from code 255 (GMY/unified)
  - Italy 1820–1861: flagged as "pre-unification, V-Dem unavailable"
  - Belgium 1820–1829: flagged as "pre-independence"
  - Belgium 1820–1845: gdppc not reconstructed in Maddison

Author: Galen Fontaise (FICSS)
Date:   May 2026
Paper:  WP-2026-006 "Koopman–CMH: Formalisation"
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings

# Suppress openpyxl warning about default style; not relevant to us
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# =============================================================================
# Configuration
# =============================================================================

PROJECT_DIR = Path(__file__).parent
RAW_DIR     = PROJECT_DIR / "raw"
PROC_DIR    = PROJECT_DIR / "processed"
PROC_DIR.mkdir(exist_ok=True)

# Input files
VDEM_FILE     = RAW_DIR / "V-Dem-CY-Full+Others-v16.csv"
MADDISON_FILE = RAW_DIR / "mpd2023_web.xlsx"
COW_FILE      = RAW_DIR / "NMC-60-abridged.csv"
PROSPECT_FILE = PROJECT_DIR / "WP-2026-006_DataCollection_Prospect_v0-1.xlsx"

# Output files
PANEL_CSV     = PROC_DIR / "WP-2026-006_WesternEurope_Panel_v0-1.csv"
PROSPECT_OUT  = PROJECT_DIR / "WP-2026-006_DataCollection_Prospect_v0-1_filled.xlsx"
LOG_FILE      = PROC_DIR / "harmonisation_log.txt"

# Panel definition
ISO3_PANEL = ['FRA', 'GBR', 'DEU', 'ITA', 'ESP', 'NLD', 'SWE', 'BEL']

# COW codes (primary, used in raw join)
COW_CODES_PRIMARY = {
    'FRA': 220, 'GBR': 200, 'ITA': 325, 'ESP': 230,
    'NLD': 210, 'SWE': 380, 'BEL': 211,
    'DEU': 255,  # Will be remapped to 260 for years 1949-1989
}
# Germany special case: FRG codes for the East-West division period
COW_CODE_GERMANY_FRG = 260   # German Federal Republic

# Sub-windows
SUB1_START, SUB1_END = 1820, 1913
SUB2_START, SUB2_END = 1945, 2020
WAR1 = set(range(1914, 1919))   # 1914-1918
WAR2 = set(range(1939, 1945))   # 1939-1944 excluded; 1945 retained as
                                  # transition/post-war year (V-E Day May 1945;
                                  # consistent with V-Dem/Maddison practice)
INTERWAR = set(range(1919, 1939))  # 1919-1938

# Variables (5)
VARIABLES = ['pop', 'gdppc', 'v2x_libdem', 'v2x_polyarchy', 'cinc']
VAR_SOURCE = {
    'pop':           'Maddison',
    'gdppc':         'Maddison',
    'v2x_libdem':    'V-Dem',
    'v2x_polyarchy': 'V-Dem',
    'cinc':          'COW',
}

# =============================================================================
# Logging
# =============================================================================

log_lines = []

def log(msg):
    timestamp = datetime.now().strftime("%H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    log_lines.append(line)

# =============================================================================
# 1. Load and slice source datasets
# =============================================================================

log("=" * 70)
log("WP-2026-006 Panel Harmonisation")
log(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
log("=" * 70)

log("\n[1/5] Loading source datasets...")

# V-Dem: load only required columns
log("  Reading V-Dem v16...")
vdem_cols = ['country_text_id', 'year', 'v2x_libdem', 'v2x_polyarchy']
vdem = pd.read_csv(VDEM_FILE, usecols=vdem_cols)
vdem = vdem[vdem['country_text_id'].isin(ISO3_PANEL)].rename(
    columns={'country_text_id': 'ISO3'})
log(f"    {len(vdem):,} rows after panel filter")

# Maddison
log("  Reading Maddison 2023...")
maddison = pd.read_excel(MADDISON_FILE, sheet_name='Full data')
maddison = maddison[maddison['countrycode'].isin(ISO3_PANEL)].rename(
    columns={'countrycode': 'ISO3'})[['ISO3', 'year', 'pop', 'gdppc']]
log(f"    {len(maddison):,} rows after panel filter")

# COW: need both 255 (unified Germany) and 260 (FRG)
log("  Reading COW NMC v6.0...")
cow_raw = pd.read_csv(COW_FILE, usecols=['ccode', 'year', 'cinc'])
cow_codes_needed = list(COW_CODES_PRIMARY.values()) + [COW_CODE_GERMANY_FRG]
cow = cow_raw[cow_raw['ccode'].isin(cow_codes_needed)].copy()
log(f"    {len(cow):,} rows after panel filter (incl. FRG code 260 for Germany)")

# =============================================================================
# 2. Build unified COW dataset with Germany special handling
# =============================================================================

log("\n[2/5] Resolving Germany code switching for COW...")

# Build a country-year COW dataset using:
#   - code 255 for non-Germany countries at all years
#   - code 255 for Germany pre-1945 and >=1990
#   - code 260 for Germany 1949-1989
#   - no data (will be NaN) for Germany 1945-1948 (occupation)

cow_panel_rows = []

for iso3, code in COW_CODES_PRIMARY.items():
    if iso3 != 'DEU':
        # Simple case: one code throughout
        sub = cow[cow['ccode'] == code][['year', 'cinc']].copy()
        sub['ISO3'] = iso3
        cow_panel_rows.append(sub[['ISO3', 'year', 'cinc']])
    else:
        # Germany: composite based on year
        deu_255 = cow[cow['ccode'] == 255][['year', 'cinc']].copy()
        deu_260 = cow[cow['ccode'] == 260][['year', 'cinc']].copy()

        # Use 255 for years <=1944 or >=1990
        deu_unified = deu_255[(deu_255['year'] <= 1944) |
                              (deu_255['year'] >= 1990)].copy()
        # Use 260 for years 1949-1989
        deu_frg = deu_260[(deu_260['year'] >= 1949) &
                          (deu_260['year'] <= 1989)].copy()

        deu_combined = pd.concat([deu_unified, deu_frg], ignore_index=True)
        deu_combined = deu_combined.sort_values('year').reset_index(drop=True)
        deu_combined['ISO3'] = 'DEU'
        cow_panel_rows.append(deu_combined[['ISO3', 'year', 'cinc']])

        log(f"    Germany: {len(deu_unified)} rows from code 255 "
            f"(pre-1945 + post-1989), "
            f"{len(deu_frg)} rows from code 260 (FRG 1949-1989)")
        log(f"    Germany 1945-1948: no data (Allied occupation)")

cow_panel = pd.concat(cow_panel_rows, ignore_index=True)
log(f"    Total COW panel rows: {len(cow_panel):,}")

# =============================================================================
# 3. Merge the three datasets
# =============================================================================

log("\n[3/5] Joining V-Dem + Maddison + COW on (ISO3, year)...")

# Outer join so we don't lose any country-year combination
merged = vdem.merge(maddison, on=['ISO3', 'year'], how='outer')
merged = merged.merge(cow_panel, on=['ISO3', 'year'], how='outer')

# Restrict to panel countries and sub-windows (drop interwar)
merged = merged[merged['ISO3'].isin(ISO3_PANEL)].copy()

def in_window(y):
    if y in WAR1 or y in WAR2 or y in INTERWAR:
        return False
    return (SUB1_START <= y <= SUB1_END) or (SUB2_START <= y <= SUB2_END)

merged = merged[merged['year'].apply(in_window)].copy()
merged = merged.sort_values(['ISO3', 'year']).reset_index(drop=True)

log(f"    Merged panel: {len(merged):,} country-year rows "
    f"(expected {8 * (94 + 76)} = 1,360)")

# Add sub-window label
merged['sub_window'] = merged['year'].apply(
    lambda y: '1820-1913' if y <= SUB1_END else '1945-2020')

# =============================================================================
# 4. Reshape to long format and produce the canonical panel CSV
# =============================================================================

log("\n[4/5] Building canonical long-format panel CSV...")

long_rows = []
for _, row in merged.iterrows():
    for var in VARIABLES:
        value = row[var] if var in row else np.nan
        long_rows.append({
            'ISO3': row['ISO3'],
            'year': int(row['year']),
            'sub_window': row['sub_window'],
            'variable': var,
            'value': value,
            'source': VAR_SOURCE[var],
        })

panel_long = pd.DataFrame(long_rows)

# Add a 'notes' column for systematic unavailability
def systematic_note(iso3, year, variable):
    """Return a note string if the (iso3, year, variable) is systematically
    unavailable for a known historical reason, else empty string."""
    notes = []
    # Italy pre-unification
    if iso3 == 'ITA' and year < 1862 and variable in ('v2x_libdem', 'v2x_polyarchy'):
        notes.append("pre-unification (Italian state from 1861); V-Dem first year 1862")
    # Belgium pre-independence
    if iso3 == 'BEL' and year < 1830:
        if variable in ('v2x_libdem', 'v2x_polyarchy', 'cinc'):
            notes.append("pre-independence (Belgium independent 1830)")
    # Belgium Maddison gdppc early
    if iso3 == 'BEL' and year < 1846 and variable == 'gdppc':
        notes.append("Maddison: gdppc not reconstructed pre-1846")
    # Germany Allied occupation
    if iso3 == 'DEU' and 1945 <= year <= 1948 and variable == 'cinc':
        notes.append("Germany under Allied occupation; no COW state coding")
    # Germany 1949-1954: FRG existed but COW does not assign cinc until full
    # sovereignty (Deutschlandvertrag, 5 May 1955)
    if iso3 == 'DEU' and 1949 <= year <= 1954 and variable == 'cinc':
        notes.append("FRG pre-sovereignty (full sovereignty 1955); COW cinc unavailable")
    # Germany 1949-1989 cinc note (Optional, info for reader)
    # We don't flag this as "systematic unavailable" because the value IS present
    # (taken from FRG code 260), but record the substitution.
    return "; ".join(notes)

panel_long['note'] = panel_long.apply(
    lambda r: systematic_note(r['ISO3'], r['year'], r['variable']),
    axis=1
)

# Save canonical CSV
panel_long.to_csv(PANEL_CSV, index=False, encoding='utf-8')
log(f"    Saved: {PANEL_CSV.name} ({len(panel_long):,} rows)")

# Coverage summary
log("\n  Coverage summary (per country, listwise complete country-years):")
wide = panel_long.pivot_table(
    index=['ISO3', 'year'], columns='variable',
    values='value', aggfunc='first'
).reset_index()
for sw_label, mask in [
    ('1820-1913', wide['year'].between(SUB1_START, SUB1_END)),
    ('1945-2020', wide['year'].between(SUB2_START, SUB2_END))
]:
    log(f"    Sub-window {sw_label}:")
    for iso3 in ISO3_PANEL:
        c = wide[(wide['ISO3'] == iso3) & mask]
        listwise = c.dropna(subset=VARIABLES)
        log(f"      {iso3}: {len(listwise):>3}/{len(c):>3} listwise complete")

# Granular DEU diagnostics for sub-window 2 (to verify special handling)
log("\n  Germany sub-window 2 granular check (which variables missing by year):")
deu_sw2 = wide[(wide['ISO3'] == 'DEU') &
               wide['year'].between(SUB2_START, SUB2_END)].sort_values('year')
for _, r in deu_sw2.iterrows():
    missing = [v for v in VARIABLES if pd.isna(r[v])]
    if missing:
        log(f"      {int(r['year'])}: missing {missing}")

# =============================================================================
# 5. Fill the prospect Excel workbook
# =============================================================================

log("\n[5/5] Populating prospect Excel workbook...")

if not PROSPECT_FILE.exists():
    log(f"  WARNING: prospect file not found at {PROSPECT_FILE}")
    log("  Skipping Excel population step.")
else:
    from openpyxl import load_workbook
    wb = load_workbook(PROSPECT_FILE)

    # Build a lookup: (ISO3, year, variable) -> (value, note)
    lookup = {}
    for _, r in panel_long.iterrows():
        lookup[(r['ISO3'], r['year'], r['variable'])] = (r['value'], r['note'])

    sheets_to_fill = ['V-Dem', 'Maddison', 'COW']
    fill_stats = {s: {'filled': 0, 'flagged': 0, 'empty': 0} for s in sheets_to_fill}

    for sheet_name in sheets_to_fill:
        if sheet_name not in wb.sheetnames:
            log(f"  WARNING: sheet '{sheet_name}' not in workbook")
            continue
        ws = wb[sheet_name]
        # Header row 1, data rows start at row 2
        # Column mapping (from prospect structure):
        #   A=ISO3, B=Country, C=Year, D=Sub_Window, E=Variable,
        #   F=Unit, G=CMH_Dim, H=Value, I=Source_Notes, J=Source_Reference
        for row_idx in range(2, ws.max_row + 1):
            iso3 = ws.cell(row=row_idx, column=1).value
            year = ws.cell(row=row_idx, column=3).value
            var  = ws.cell(row=row_idx, column=5).value
            if iso3 is None or year is None or var is None:
                continue
            key = (iso3, int(year), var)
            if key in lookup:
                value, note = lookup[key]
                if pd.notna(value):
                    ws.cell(row=row_idx, column=8).value = float(value)
                    fill_stats[sheet_name]['filled'] += 1
                else:
                    fill_stats[sheet_name]['empty'] += 1
                if note:
                    ws.cell(row=row_idx, column=9).value = note
                    fill_stats[sheet_name]['flagged'] += 1

    for sheet_name in sheets_to_fill:
        s = fill_stats[sheet_name]
        log(f"  {sheet_name}: {s['filled']} values filled, "
            f"{s['empty']} empty (missing), {s['flagged']} flagged in Source_Notes")

    wb.save(PROSPECT_OUT)
    log(f"  Saved: {PROSPECT_OUT.name}")

# =============================================================================
# Final report
# =============================================================================

log("\n" + "=" * 70)
log("HARMONISATION COMPLETE")
log("=" * 70)
log(f"  Canonical panel CSV: {PANEL_CSV}")
log(f"  Filled prospect:     {PROSPECT_OUT}")
log(f"  Log:                 {LOG_FILE}")

# Save log
with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write("\n".join(log_lines))
